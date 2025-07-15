import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import calendar
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
import csv


# Colores 
colores = ['#FFB897', '#B8E6A7', '#809bce', "#64a09d", '#CBE6FF', '#E6E6FA']
# Meses 
meses_en_espanol = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 
    7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

# -------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------- COMPARATIVAS POR AÑO - HOJA: COMPARATIVA AÑO DTO/PCL  --------------------------------------
# -------------------------------------------------------------------------------------------------------------------------------------------------------------------
def graficas_barras_tabla_mes_comparativa(df, nombre_hoja):
    # Filtrar solo los datos de BELISARIO397 y GESTAR INNOVACION
    df_comparativa = df[df['NOTIFICADOR'].isin(['BELISARIO 397', 'GESTAR INNOVACION'])]
    
    conteo = df_comparativa.groupby(['MES', 'NOTIFICADOR']).size().unstack(fill_value=0)
    conteo.index = conteo.index.map(lambda m: meses_en_espanol[m].capitalize())

    # Crear la gráfica de barras
    fig, ax = plt.subplots(figsize=(12, 8))
    conteo.plot(kind='bar', ax=ax, color=colores)
    ax.set_xlabel('Mes')
    ax.set_ylabel('Número de Datos')
    ax.legend(title='Notificadores', bbox_to_anchor=(1.2, 1), loc='upper left', fontsize=10)

    for p in ax.patches:
        ax.annotate(f'{p.get_height()}', 
                    (p.get_x() + p.get_width() / 2., p.get_height()), 
                    xytext=(0, 5),
                    textcoords='offset points',
                    ha='center', va='bottom', fontsize=10, color='black')

    grafico_path = f"{nombre_hoja}_grafico_barras_comparativa.png"
    plt.tight_layout()
    plt.savefig(grafico_path, transparent=True, bbox_inches="tight")
    return grafico_path

def graficapastel_comparativa_ano(df, nombre_hoja):
    # Filtrar solo los datos de BELISARIO397 y GESTAR INNOVACION
    df_comparativa = df[df['NOTIFICADOR'].isin(['BELISARIO 397', 'GESTAR INNOVACION'])]
    
    # Crear gráfico de pastel comparativo por notificadores
    conteo = df_comparativa.groupby('NOTIFICADOR').size()
    
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.pie(conteo, labels=None, autopct='%1.1f%%', startangle=90, colors=colores)
    ax.legend(labels=conteo.index, title='Notificadores', loc='center left', bbox_to_anchor=(1.05, 0.5), fontsize=10)

    # Guardar el gráfico en un archivo
    grafico_path = f"{nombre_hoja}_grafico_pastel_ano.png"
    plt.tight_layout()
    plt.savefig(grafico_path, transparent=True, bbox_inches="tight")
    return grafico_path
# ---------------------------------------------------------------------- TABLAS  --------------------------------------

def tabla_comparativa_por_mes(df, hoja):
    # Filtrar solo los datos de interés
    df_comparativa = df[df['NOTIFICADOR'].isin(['BELISARIO 397', 'GESTAR INNOVACION'])]

    # Agrupar y pivotear
    conteo = df_comparativa.groupby(['MES', 'NOTIFICADOR']).size().unstack(fill_value=0)
    conteo.index = conteo.index.map(lambda m: meses_en_espanol[m].capitalize())
    conteo.index.name = "MES"

    # Preparar estilos
    borde_fino = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for r_idx, fila in enumerate(dataframe_to_rows(conteo.reset_index(), index=False, header=True), start=1):
        for c_idx, valor in enumerate(fila, start=1):
            celda = hoja.cell(row=r_idx, column=c_idx, value=valor)

            # Estilo general
            celda.alignment = Alignment(horizontal="center")
            celda.border = borde_fino

            # Negrita para headers
            if r_idx == 1:
                celda.font = Font(bold=True)

# ---------------------------------------------------------------------- Hojas  --------------------------------------

def crear_comparativa_ano_dto(libro, df_dto):
    if "COMPARATIVA AÑO DTO" in libro.sheetnames:
        del libro["COMPARATIVA AÑO DTO"]
    hoja = libro.create_sheet("COMPARATIVA AÑO DTO")

    df_comparativa = df_dto[df_dto['NOTIFICADOR'].isin(['BELISARIO 397', 'GESTAR INNOVACION'])]

    # 👉 Primero la tabla
    tabla_comparativa_por_mes(df_comparativa, hoja)

    # Luego los gráficos (en posiciones fijas que no pisen la tabla)
    grafico_barras_comparativa_path = graficas_barras_tabla_mes_comparativa(df_comparativa, "COMPARATIVA AÑO DTO")
    hoja.add_image(Image(grafico_barras_comparativa_path), 'I4')

    grafico_pastel_comparativa_path = graficapastel_comparativa_ano(df_comparativa, "COMPARATIVA AÑO DTO")
    hoja.add_image(Image(grafico_pastel_comparativa_path), 'I4')

# Hoja "COMPARATIVA AÑO PCL"
def crear_comparativa_ano_pcl(libro, df_pcl):
    if "COMPARATIVA AÑO PCL" in libro.sheetnames:
        del libro["COMPARATIVA AÑO PCL"]
    hoja = libro.create_sheet("COMPARATIVA AÑO PCL")

    df_comparativa = df_pcl[df_pcl['NOTIFICADOR'].isin(['BELISARIO 397', 'GESTAR INNOVACION'])]

    # 👉 Primero la tabla
    tabla_comparativa_por_mes(df_comparativa, hoja)

    # Luego los gráficos en otra parte de la hoja
    grafico_barras_comparativa_path = graficas_barras_tabla_mes_comparativa(df_comparativa, "COMPARATIVA AÑO PCL")
    hoja.add_image(Image(grafico_barras_comparativa_path), 'I4')

    grafico_pastel_comparativa_path = graficapastel_comparativa_ano(df_comparativa, "COMPARATIVA AÑO PCL")
    hoja.add_image(Image(grafico_pastel_comparativa_path), 'I4')




# -------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------  HOJA MES PCL/DTO_MES   ---------------------------------------------------------------------
# -------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------- GRAFICOS ----------------------

def graficas_barras_hojames(df, nombre_hoja, mes):
    df_filtrado = df[df['MES'] == mes]
    conteo = df_filtrado.groupby(['NOTIFICADOR', 'ESTADO_INFORME']).size().unstack(fill_value=0)

    fig, ax = plt.subplots(figsize=(14, 8))
    conteo.plot(kind='bar', ax=ax, color=colores[:len(conteo.columns)])
    ax.set_xlabel('Notificador')
    ax.set_ylabel('Cantidad')
    ax.legend(title='Estado Informe', bbox_to_anchor=(1.2, 1), loc='upper left', fontsize=10)

    for p in ax.patches:
        height = p.get_height()
        if height > 0:
            ax.annotate(f'{int(height)}',
                        (p.get_x() + p.get_width() / 2., height),
                        xytext=(0, 5),
                        textcoords='offset points',
                        ha='center', va='bottom', fontsize=9, color='black')

    grafico_path = f"{nombre_hoja}_grafico_barras_hojames_{mes}.png"
    plt.tight_layout()
    plt.savefig(grafico_path, transparent=True, bbox_inches="tight")
    plt.close(fig)
    return grafico_path

def graficas_pastel_hoja_mes(df, nombre_hoja, mes):
    if 'MES' not in df.columns:
        df['MES'] = df['FECHA_VISADO'].dt.month

    df_mes = df[df['MES'] == mes]
    conteo = (
        df_mes
        .groupby(['NOTIFICADOR', 'ESTADO_INFORME'])
        .size()
        .reset_index(name='CUENTA')
    )

    conteo['ETIQUETA'] = conteo['NOTIFICADOR'] + " – " + conteo['ESTADO_INFORME']
    fig, ax = plt.subplots(figsize=(10, 8))
    cmap = plt.cm.get_cmap('tab20', len(conteo))

    wedges, _, _ = ax.pie(
        conteo['CUENTA'],
        labels=None,
        autopct='%1.1f%%',
        startangle=140,
        colors=[cmap(i) for i in range(len(conteo))]
    )

    ax.axis('equal')
    ax.legend(
        wedges,
        conteo['ETIQUETA'],
        title="Notificador – Estado",
        loc='center left',
        bbox_to_anchor=(1, 0.5),
        fontsize=9
    )

    path = f"{nombre_hoja}_pastel_notificador_estado_{mes}.png"
    plt.tight_layout()
    plt.savefig(path, transparent=True, bbox_inches="tight")
    plt.close(fig)
    return path
# ---------------------- HOJA SOLO DATOS ----------------------

def crear_hoja_datos_mes(libro, df_mes, tipo, mes):
    mes_nombre = meses_en_espanol.get(mes, f"Mes{mes}")
    nombre_hoja = f"{tipo}_{mes_nombre}"

    if nombre_hoja in libro.sheetnames:
        del libro[nombre_hoja]
    hoja = libro.create_sheet(nombre_hoja)

    for r, fila in enumerate(dataframe_to_rows(df_mes, index=False, header=True), start=1):
        for c, valor in enumerate(fila, start=1):
            hoja.cell(row=r, column=c, value=valor)

    return nombre_hoja

# ---------------------- HOJA CON GRAFICOS + TABLA ----------------------
def resumen_notificador_estado(df, mes):
    df_mes = df[df['MES'] == mes]
    resumen = (
        df_mes
        .groupby(['NOTIFICADOR', 'ESTADO_INFORME'])
        .size()
        .reset_index(name='TOTAL')  
        .sort_values(by=['NOTIFICADOR', 'ESTADO_INFORME'])
    )
    return resumen


def tabla_hojames(libro, df_mes, tipo, mes, pos_tabla='A1', pos_barras='H5', pos_pastel='H35'):
    mes_nombre = meses_en_espanol.get(mes, f"Mes{mes}")
    nombre_hoja = f"{tipo}_{mes_nombre}_tabla_graficos"

    if 'MES' not in df_mes.columns:
        df_mes['MES'] = df_mes['FECHA_VISADO'].dt.month

    if nombre_hoja in libro.sheetnames:
        del libro[nombre_hoja]
    hoja = libro.create_sheet(nombre_hoja)

    # 🧮 Crear resumen de datos
    resumen = resumen_notificador_estado(df_mes, mes)

    # 💾 Volcar tabla resumida con estilos
    rows = list(dataframe_to_rows(resumen, index=False, header=True))

    # 🎨 Definir estilo
    borde = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    fill_gris = PatternFill("solid", fgColor="D9D9D9")
    font_bold = Font(bold=True)

    for r_idx, fila in enumerate(rows, start=1):
        for c_idx, valor in enumerate(fila, start=1):
            celda = hoja.cell(row=r_idx, column=c_idx, value=valor)
            celda.border = borde
            if r_idx == 1:
                celda.fill = fill_gris
                celda.font = font_bold

    # 📊 Agregar gráficos
    barra_path = graficas_barras_hojames(df_mes, nombre_hoja, mes)
    hoja.add_image(Image(barra_path), pos_barras)

    pastel_path = graficas_pastel_hoja_mes(df_mes, nombre_hoja, mes)
    hoja.add_image(Image(pastel_path), pos_pastel)

    return nombre_hoja


# ---------------------- CREA AMBAS HOJAS DTO/PCL ----------------------

def crear_hojas_dto_pcl_tabla(libro, df_total, mes):
    if 'MES' not in df_total.columns:
        df_total['MES'] = df_total['FECHA_VISADO'].dt.month

    if 'HOJA_ORIGEN' not in df_total.columns:
        raise ValueError("Falta la columna HOJA_ORIGEN (debe valer 'DTO' o 'PCL').")

    filtros = {
        'PCL': df_total[(df_total['HOJA_ORIGEN'] == 'PCL') & (df_total['MES'] == mes)],
        'DTO': df_total[(df_total['HOJA_ORIGEN'] == 'DTO') & (df_total['MES'] == mes)]
    }

    for tipo, df_mes in filtros.items():
        if df_mes.empty:
            print(f"⚠️ Nada para {tipo} en el mes {mes}, se salta la hoja.")
            continue

        crear_hoja_datos_mes(libro, df_mes, tipo, mes)   # 📄 Datos con nombre bonito
        tabla_hojames(libro, df_mes, tipo, mes)          # 📊 Tabla y gráficos

    print("✅ Hojas DTO/PCL del mes creadas: datos + gráficos 🎉")
# -------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------  HOJA: TABLA MES  ---------------------------------------------------------------------
# -------------------------------------------------------------------------------------------------------------------------------------------------------------------
def _mes_a_nombre(m):
    if isinstance(m, str):
        return m.capitalize()
    return meses_en_espanol[int(m)].capitalize()

def graficas_barras_tabla_mes(df, nombre_hoja):
    if 'MES' not in df.columns:
        df['MES'] = df['FECHA_VISADO'].dt.month
    df['MES_NOMBRE'] = df['MES'].apply(_mes_a_nombre)

    # Agrupar por MES_NOMBRE y NOTIFICADOR → contar casos
    conteo = df.groupby(['MES_NOMBRE', 'NOTIFICADOR']).size().reset_index(name='TOTAL')

    # Pivot sin inventar meses
    tabla = conteo.pivot(index='MES_NOMBRE', columns='NOTIFICADOR', values='TOTAL').fillna(0)

    # Ordenar meses sin crear filas nuevas
    orden_meses = [_mes_a_nombre(i) for i in range(1, 13)]
    tabla.index = pd.CategoricalIndex(tabla.index, categories=orden_meses, ordered=True)
    tabla = tabla.sort_index()

    # 📊 Plot
    fig, ax = plt.subplots(figsize=(12, 6))
    tabla.plot(kind='bar', ax=ax, color=colores)

    ax.set_xlabel("Mes")
    ax.set_ylabel("Cantidad")
    plt.xticks(rotation=45, ha='right')
    ax.legend(title='Notificador', bbox_to_anchor=(1.05, 1), loc='upper left')

    path = f"{nombre_hoja}_grafico_barras_mensual.png"
    plt.tight_layout()
    plt.savefig(path, transparent=True, bbox_inches="tight")
    plt.close(fig)
    return path


def graficas_pastel_tabla_mes(df, nombre_hoja):
    if 'MES' not in df.columns:
        df['MES'] = df['FECHA_VISADO'].dt.month

    # Convertir al vuelo sin tocar df
    conteo = (
        df.groupby('MES').size()
          .rename(index=_mes_a_nombre)
    )

    fig, ax = plt.subplots(figsize=(8, 8))
    cmap = plt.cm.get_cmap('Pastel1', len(conteo))
    ax.pie(conteo, labels=None, autopct='%1.1f%%',
           startangle=90, colors=[cmap(i) for i in range(len(conteo))])
    ax.legend(labels=conteo.index, title='Meses', loc='center left',
              bbox_to_anchor=(1.05, 0.5), fontsize=10)

    path = f"{nombre_hoja}_grafico_pastel.png"
    plt.tight_layout(); plt.savefig(path, transparent=True,
                                    bbox_inches="tight"); plt.close(fig)
    return path

def grafica_pastel_tabla_mes_porproveedor(df, nombre_hoja):
    proveedores = df['NOTIFICADOR'].dropna().unique()
    rutas = []

    for proveedor in proveedores:
        df_prov = df[df['NOTIFICADOR'] == proveedor]
        conteo = df_prov['ESTADO_INFORME'].value_counts()

        if conteo.empty:
            continue  # Nada que graficar

        fig, ax = plt.subplots(figsize=(8, 8))
        cmap = plt.cm.get_cmap('Pastel2', len(conteo))
        ax.pie(
            conteo,
            labels=None,
            autopct='%1.1f%%',
            startangle=90,
            colors=[cmap(i) for i in range(len(conteo))]
        )

        ax.set_title(f"{proveedor}", fontsize=12)
        ax.legend(labels=conteo.index, title='Estado Informe',
                  loc='center left', bbox_to_anchor=(1.05, 0.5), fontsize=10)

        filename = f"{nombre_hoja}_pastel_{proveedor.replace(' ', '_').replace('/', '-')}.png"
        plt.tight_layout()
        plt.savefig(filename, transparent=True, bbox_inches="tight")
        plt.close(fig)

        rutas.append(filename)

    return rutas


# ------------------------------------------------------------------------------- GENERAR TABLAS PARA DTO Y PCL: TABLA MES -------------------------------------------------------------
def generar_tablas_dto_y_pcl(libro, df_dto, df_pcl):
    def crear_hoja(nombre_hoja, df):
        df['MES'] = df['FECHA_VISADO'].dt.month

        conteo = df.groupby('MES').size().reset_index(name='TOTAL')
        conteo['MES'] = conteo['MES'].map(lambda m: meses_en_espanol[m].capitalize())

        total_general = conteo['TOTAL'].sum()
        conteo['PORCENTAJE'] = (
            (conteo['TOTAL'] / total_general * 100)
            .round(2).astype(str) + '%'
        )

        fila_total = pd.DataFrame({
            'MES': ['Total general'],
            'TOTAL': [total_general],
            'PORCENTAJE': ['100.0%']
        })

        tabla_final = pd.concat([conteo, fila_total], ignore_index=True)

        # Crear hoja
        if nombre_hoja in libro.sheetnames:
            del libro[nombre_hoja]
        hoja = libro.create_sheet(nombre_hoja)

        # Definir estilos
        borde_oscuro = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        relleno_encabezado = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        alineacion_centrada = Alignment(horizontal='center', vertical='center')
        fuente_negrita = Font(bold=True)

        # Escribir encabezados con estilo
        for c_idx, col_name in enumerate(tabla_final.columns, 1):
            celda = hoja.cell(row=1, column=c_idx, value=col_name)
            celda.fill = relleno_encabezado
            celda.border = borde_oscuro
            celda.alignment = alineacion_centrada
            celda.font = fuente_negrita

        # Escribir datos con bordes
        for r_idx, row in tabla_final.iterrows():
            for c_idx, value in enumerate(row, 1):
                celda = hoja.cell(row=r_idx+2, column=c_idx, value=value)
                celda.border = borde_oscuro
                celda.alignment = alineacion_centrada

        # Gráficos
        grafico_barras_path = graficas_barras_tabla_mes(df, nombre_hoja)
        hoja.add_image(Image(grafico_barras_path), 'E5')

        grafico_pastel_path = graficas_pastel_tabla_mes(df, nombre_hoja)
        hoja.add_image(Image(grafico_pastel_path), 'E20')

        grafico_pastel_proveedor_path = grafica_pastel_tabla_mes_porproveedor(df, nombre_hoja)
        hoja.add_image(Image(grafico_pastel_proveedor_path[0]), 'E35')

    crear_hoja("DTO TABLA MES", df_dto)
    crear_hoja("PCL TABLA MES", df_pcl)



# -------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------  FUNCIONES DE SUBIDA Y DESCARGA  ---------------------------------------------------------------------
# -------------------------------------------------------------------------------------------------------------------------------------------------------------------

def descargar_archivo(output, nombre="archivo_procesado.xlsx"):
    st.download_button(
        label="📥 Descargar archivo",  
        data=output,
        file_name=nombre,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



def subir_archivo():
    archivo = st.file_uploader("Sube un archivo (.xlsx o .csv)", type=["xlsx", "csv"])

    if archivo is not None:
        try:
            nombre_archivo = archivo.name.lower()

            if nombre_archivo.endswith(".xlsx"):
                xls = pd.ExcelFile(archivo)
                hojas = xls.sheet_names

                if "DTO" in hojas and "PCL" in hojas:
                    st.success("¡Archivo Excel válido! Se encontraron las hojas DTO y PCL.")
                    return archivo, "xlsx"
                else:
                    if "DTO" not in hojas:
                        st.error("La hoja 'DTO' no se encuentra en el archivo.")
                    if "PCL" not in hojas:
                        st.error("La hoja 'PCL' no se encuentra en el archivo.")
                    return None, None

            elif nombre_archivo.endswith(".csv"):
                df = pd.read_csv(archivo)
                if "DTO" in df.columns and "PCL" in df.columns:
                    st.success("¡Archivo CSV válido! Se encontraron las columnas DTO y PCL.")
                    return archivo, "csv"
                else:
                    st.warning("El archivo CSV no contiene columnas llamadas 'DTO' y 'PCL'.")
                    return None, None

        except Exception as e:
            st.error(f"Error al procesar el archivo: {e}")
            return None, None

    return None, None

# ------------------------------------------------------------------------------- FLUJO ---------------------------------------------------------------------------------
def procesar_archivos():
    archivo, tipo = subir_archivo()

    if archivo and tipo == "xlsx":
        # Mostrar el selector de mes con los meses en español
        mes_seleccionado = st.selectbox("Selecciona el mes", list(meses_en_espanol.values()))  # Ahora muestra los meses en español

        # Leer las hojas DTO y PCL
        df_dto = pd.read_excel(archivo, sheet_name='DTO', parse_dates=['FECHA_VISADO'])
        df_pcl = pd.read_excel(archivo, sheet_name='PCL', parse_dates=['FECHA_VISADO'])

        # Crear archivo con los datos filtrados por el mes seleccionado
        archivo.seek(0)
        archivo_bytes = BytesIO(archivo.read())
        libro = load_workbook(archivo_bytes)

        # Convertir el mes seleccionado a número usando el diccionario
        mes_num = list(meses_en_espanol.values()).index(mes_seleccionado) + 1  # Obtiene el índice del mes (1-12)
        # 👇🏽 AÑADE ESTO AQUÍ
        df_dto_mes = df_dto[df_dto['FECHA_VISADO'].dt.month == mes_num]
        df_pcl_mes = df_pcl[df_pcl['FECHA_VISADO'].dt.month == mes_num]


        # Llamar a la función para generar las hojas con el mes seleccionado    
        crear_hoja_datos_mes(libro, df_dto, "DTO", mes_num)
        tabla_hojames(libro, df_dto, "DTO", mes_num)
        crear_hoja_datos_mes(libro, df_pcl, "PCL", mes_num)
        tabla_hojames(libro, df_pcl, "PCL", mes_num)


        # Llamar a la función para generar las tablas de DTO y PCL
        generar_tablas_dto_y_pcl(libro, df_dto, df_pcl)

        # Llamar a la función para crear la hoja de comparativa de año
        crear_comparativa_ano_dto(libro, df_dto)
        crear_comparativa_ano_pcl(libro, df_pcl)
        
        # Llamar a la función para crear las tablas de HOJA MES
        tabla_hojames(libro, df_dto_mes, 'DTO', mes_num)
        tabla_hojames(libro, df_pcl_mes, 'PCL', mes_num)


        output = BytesIO()
        libro.save(output)
        output.seek(0)
        descargar_archivo(output, nombre="informe_dto_pcl_mes.xlsx")
        st.success("✅ Archivo generado con éxito.")
    elif archivo and tipo == "csv":
        st.warning("Actualmente el procesamiento está disponible solo para archivos .xlsx con hojas DTO y PCL.")
